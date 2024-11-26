import json
import boto3
import requests
import os
import openpyxl
import io
import base64

from decimal import Decimal
from uuid import uuid4
from datetime import datetime
from decouple import config

# Configuración DynamoDB
AWS_REGION = "us-east-1"
DYNAMO_TABLE = config("DYNAMO_TABLE")
PRODUCTS_API_URL = config("PRODUCTS_API_URL")
BUCKET_NAME = os.environ['S3_BUCKET_NAME']


dynamodb = boto3.resource("dynamodb", region_name=AWS_REGION)
table = dynamodb.Table(DYNAMO_TABLE)


def decimal_to_serializable(obj):
    if isinstance(obj, Decimal):
        return float(obj) if obj % 1 else int(obj)
    if isinstance(obj, list):
        return [decimal_to_serializable(i) for i in obj]
    if isinstance(obj, dict):
        return {k: decimal_to_serializable(v) for k, v in obj.items()}
    return obj


def create_inbound_note(event, context):
    body = json.loads(event["body"])
    
    # Define required fields and their expected types
    required_fields = {
        "Date": str,
        "Products": list
    }
    
    errors = []
    
    # Validate required fields and their types
    for field, expected_type in required_fields.items():
        if field not in body:
            errors.append(f"Field '{field}' is required.")
        else:
            if not isinstance(body[field], expected_type):
                errors.append(f"Field '{field}' must be of type {expected_type.__name__}.")

    # Return errors if any validations fail
    if errors:
        return {
            "statusCode": 400,
            "headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        },
            "body": json.dumps({"errors": errors}),
        }
    
    note_id = str(uuid4())
    date = body["Date"]
    products = body["Products"]
    
    # Validate format and contents of products
    if not products:
        return {
            "statusCode": 400,
            "headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        },
            "body": json.dumps({"message": "At least one product is required."}),
        }

    product_errors = []
    for idx, product in enumerate(products):
        if not isinstance(product, dict):
            product_errors.append(f"Product at index {idx} must be an object.")
            continue

        # Required fields in each product
        product_required_fields = {
            "ProductID": str,
            "Quantity": int
        }

        for field, expected_type in product_required_fields.items():
            if field not in product:
                product_errors.append(f"Field '{field}' is required in product at index {idx}.")
            else:
                if not isinstance(product[field], expected_type):
                    product_errors.append(
                        f"Field '{field}' in product at index {idx} must be of type {expected_type.__name__}."
                    )
                elif field == "Quantity" and product["Quantity"] <= 0:
                    product_errors.append(
                        f"Invalid quantity for product '{product.get('ProductID', 'unknown')}' at index {idx}. Quantity must be greater than 0."
                    )

    # Return errors if any product validations fail
    if product_errors:
        return {
            "statusCode": 400,
            "headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        },
            "body": json.dumps({"errors": product_errors}),
        }

    # Call the products API to update each product
    for product in products:
        response = requests.put(
            f"{PRODUCTS_API_URL}/{product['ProductID']}",
            json={"Quantity": product["Quantity"]},
        )
        if response.status_code != 200:
            return {
                "statusCode": 400,
                "headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        },
                "body": json.dumps(
                    {
                        "message": f"Failed to update product {product['ProductID']}: {response.text}"
                    }
                ),
            }

    # Save the note in DynamoDB
    note = {"NoteID": note_id, "Date": date, "Products": products}
    table.put_item(Item=note)

    return {
        "statusCode": 201,
        "headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        },
        "body": json.dumps({"message": "Inbound note created", "NoteID": note_id}),
    }


def update_inbound_note(event, context):
    note_id = event["pathParameters"]["note_id"]

    # Fetch the current note from DynamoDB
    response = table.get_item(Key={"NoteID": note_id})
    if "Item" not in response:
        return {
            "statusCode": 404,
            "headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        },
            "body": json.dumps({"message": "Note not found"}),
        }
    current_note = response["Item"]

    body = json.loads(event["body"])

    if "Products" not in body or not isinstance(body["Products"], list):
        return {
            "statusCode": 400,
            "headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        },
            "body": json.dumps({"message": "Field 'Products' is required and must be a list."}),
        }

    new_products = body["Products"]

    if not new_products:
        return {
            "statusCode": 400,
            "headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        },
            "body": json.dumps({"message": "At least one product is required."}),
        }

    product_errors = []
    for idx, product in enumerate(new_products):
        if not isinstance(product, dict):
            product_errors.append(f"Product at index {idx} must be an object.")
            continue

        # Required fields in each product
        product_required_fields = {
            "ProductID": str,
            "Quantity": int
        }

        for field, expected_type in product_required_fields.items():
            if field not in product:
                product_errors.append(f"Field '{field}' is required in product at index {idx}.")
            else:
                if not isinstance(product[field], expected_type):
                    product_errors.append(
                        f"Field '{field}' in product at index {idx} must be of type {expected_type.__name__}."
                    )
                elif field == "Quantity" and product["Quantity"] <= 0:
                    product_errors.append(
                        f"Invalid quantity for product '{product.get('ProductID', 'unknown')}' at index {idx}. Quantity must be greater than 0."
                    )

    if product_errors:
        return {
            "statusCode": 400,
            "headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        },
            "body": json.dumps({"errors": product_errors}),
        }

    old_products = {p["ProductID"]: p for p in current_note.get("Products", [])}
    new_products_dict = {p["ProductID"]: p for p in new_products}

    all_product_ids = set(old_products.keys()).union(new_products_dict.keys())

    for product_id in all_product_ids:
        old_quantity = old_products.get(product_id, {}).get("Quantity", 0)
        new_quantity = new_products_dict.get(product_id, {}).get("Quantity", 0)
        quantity_diff = new_quantity - old_quantity

        if quantity_diff != 0:
            response = requests.put(
                f"{PRODUCTS_API_URL}/{product_id}",
                json={"Quantity": quantity_diff},
            )
            if response.status_code != 200:
                return {
                    "statusCode": 400,
                    "headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        },
                    "body": json.dumps(
                        {
                            "message": f"Failed to update product {product_id}: {response.text}"
                        }
                    ),
                }

    update_expression = "SET "
    expression_attribute_names = {}
    expression_attribute_values = {}

    update_expression += "#Products = :Products"
    expression_attribute_names["#Products"] = "Products"
    expression_attribute_values[":Products"] = new_products

    if "Date" in body:
        if not isinstance(body["Date"], str):
            return {
                "statusCode": 400,
                "headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        },
                "body": json.dumps({"message": "Field 'Date' must be a string."}),
            }
        update_expression += ", #Date = :Date"
        expression_attribute_names["#Date"] = "Date"
        expression_attribute_values[":Date"] = body["Date"]

    table.update_item(
        Key={"NoteID": note_id},
        UpdateExpression=update_expression,
        ExpressionAttributeValues=expression_attribute_values,
        ExpressionAttributeNames=expression_attribute_names,
    )

    return {
        "statusCode": 200,
        "headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        },
        "body": json.dumps({"message": "Inbound note updated"}),
    }


def get_all_inbound_notes(event, context):
    response = table.scan()
    items = decimal_to_serializable(response["Items"])
    return {"statusCode": 200, "headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        },"body": json.dumps(items)}


def get_inbound_note(event, context):
    note_id = event["pathParameters"]["note_id"]
    response = table.get_item(Key={"NoteID": note_id})
    if "Item" not in response:
        return {"statusCode": 404, "headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        },"body": json.dumps({"message": "Note not found"})}
    note = decimal_to_serializable(response["Item"])
    return {"statusCode": 200,"headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        }, "body": json.dumps(note)}


def delete_inbound_note(event, context):
    note_id = event["pathParameters"]["note_id"]

    response = table.get_item(Key={"NoteID": note_id})
    if "Item" not in response:
        return {"statusCode": 404,"headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        }, "body": json.dumps({"message": "Note not found"})}

    note = response["Item"]

    # Llamar al endpoint de productos para revertir las cantidades
    for product in note["Products"]:
        response = requests.put(
            f"{PRODUCTS_API_URL}/{product['ProductID']}",
            json={"Quantity": -product["Quantity"]},
        )
        if response.status_code != 200:
            return {
                "statusCode": 400,
                "headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        },
                "body": json.dumps({"message": f"Failed to revert product {product['ProductID']}: {response.text}"}),
            }

    # Eliminar la nota de DynamoDB
    table.delete_item(Key={"NoteID": note_id})
    return {"statusCode": 200, "headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        },"body": json.dumps({"message": "Inbound note deleted"})}

def get_inbound_note_file(event, context):
    # Validate that 'note_id' is provided in pathParameters
    if 'pathParameters' not in event or not event['pathParameters'] or 'note_id' not in event['pathParameters']:
        return {
            "statusCode": 400,
            "headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        },
            "body": json.dumps({"message": "The 'note_id' parameter is required."}),
        }

    note_id = event["pathParameters"]["note_id"]

    # Fetch the note from DynamoDB
    response = table.get_item(Key={"NoteID": note_id})

    if "Item" not in response:
        return {
            "statusCode": 404,
            "headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        },
            "body": json.dumps({"message": "Note not found."}),
        }

    note = decimal_to_serializable(response["Item"])

    # Create an Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Note Details'

    # Write NoteID and Date
    ws['A1'] = 'NoteID'
    ws['B1'] = note.get('NoteID')
    ws['A2'] = 'Date'
    ws['B2'] = note.get('Date')

    # Leave a blank row
    ws.append([])

    # Write product headers
    ws.append(['ProductID', 'Quantity'])

    # Write product data
    for product in note.get('Products', []):
        ws.append([product.get('ProductID'), product.get('Quantity')])

    # Save the workbook to a bytes buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Upload the Excel file to S3
    s3 = boto3.client('s3')
    object_key = f"nota_{note_id}.xlsx"

    try:
        s3.upload_fileobj(output, BUCKET_NAME, object_key)
    except Exception as e:
        return {
            "statusCode": 500,
            "headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        },
            "body": json.dumps({"message": f"Failed to upload file to S3: {str(e)}"}),
        }

    output.close()

    try:
        url = s3.generate_presigned_url(
            ClientMethod='get_object',
            Params={'Bucket': BUCKET_NAME, 'Key': object_key},
            ExpiresIn=3600 
        )
    except Exception as e:
        return {
            "statusCode": 500,
            "headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        },
            "body": json.dumps({"message": f"Failed to generate pre-signed URL: {str(e)}"}),
        }

    return {
        "statusCode": 200,
        "headers": {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token",
            "Access-Control-Allow-Methods": "OPTIONS,GET,POST,PUT,DELETE"
        },
        "body": json.dumps({"download_url": url}),
        "headers": {
            "Content-Type": "application/json"
        }
    }
